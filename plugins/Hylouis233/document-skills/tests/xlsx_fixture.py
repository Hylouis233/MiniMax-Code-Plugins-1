# Minimal runnable fixtures for the snippets called out in review: one file per format.
# Each script is self-contained, writes only scratch files into the current directory,
# and exits non-zero on failed assertions. Run from any scratch directory:
#   python pdf_fixture.py   (deps: reportlab, pypdf, pymupdf)
#   python pptx_fixture.py  (deps: python-pptx)
#   python xlsx_fixture.py  (deps: openpyxl)
#   python docx_fixture.py  (deps: python-docx, pymupdf, soffice on PATH)
import base64
import csv
import sys
import zipfile
from tempfile import TemporaryFile

import openpyxl

failures = []


def check(name, cond, extra=""):
    print(("PASS " if cond else "FAIL ") + name + ((" :: " + str(extra)) if not cond and extra else ""))
    if not cond:
        failures.append(name)


# ---- csv.md snippet: sniffed dialect actually reaches the reader ---------------
semicolon_csv = "region;units;note\nEU;120;first\nUS;80;second\n"
with open("input.csv", "w", newline="", encoding="utf-8") as f:
    f.write(semicolon_csv)

with open("input.csv", newline="", encoding="utf-8-sig") as f:
    sample = f.read(2048)
    f.seek(0)
    dialect = csv.Sniffer().sniff(sample)
    reader = csv.DictReader(f, dialect=dialect)
    rows = list(reader)

check("sniffer detects the semicolon dialect", dialect.delimiter == ";", dialect.delimiter)
check("rows parse into per-column values", rows[0] == {"region": "EU", "units": "120", "note": "first"}, rows[0])

# negative control: default comma reader splits the whole line as one key
with open("input.csv", newline="", encoding="utf-8-sig") as f:
    naive = next(csv.DictReader(f))
check("default reader is proven wrong here (negative control)", list(naive.keys())[0] == "region;units;note", naive)

# ---- csv.md conversion: formula-looking input remains literal text -----------
formula_looking = '=HYPERLINK("https://example.invalid", "click")'
csv_wb = openpyxl.Workbook()
csv_cell = csv_wb.active["A1"]
csv_cell.value = formula_looking
csv_cell.data_type = "s"
csv_wb.save("csv-text.xlsx")
csv_reopened = openpyxl.load_workbook("csv-text.xlsx", data_only=False)
check("formula-looking CSV field keeps its exact text", csv_reopened.active["A1"].value == formula_looking)
check("formula-looking CSV field is not an XLSX formula", csv_reopened.active["A1"].data_type == "s")

unsafe_wb = openpyxl.Workbook()
unsafe_wb.active["A1"] = formula_looking
check("plain assignment is proven unsafe (negative control)", unsafe_wb.active["A1"].data_type == "f")

# ---- edit.md snippet: round_trip_changes detects dropped parts AND stripped extensions ----

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"
ws.append(["Region", "Units"])
ws.append(["EU", 120])
wb.save("plain.xlsx")

with zipfile.ZipFile("plain.xlsx") as zin:
    payload = {name: zin.read(name) for name in zin.namelist()}

# simulate an unsupported extension part (what a slicer/queries part looks like in the zip)
payload["xl/slicers/slicer1.xml"] = b"<slicer xmlns='stub'/>"
payload["xl/media/large.bin"] = b"x14:" * 32_768  # binary payload must never be marker-scanned

# simulate an in-part x14 extension that openpyxl will strip while keeping the part name
EXT = (b"<extLst><ext uri='{00000000-0000-0000-0000-000000000000}' "
       b"xmlns:x14='http://schemas.microsoft.com/office/spreadsheetml/2009/9/main'>"
       b"<x14:stub/></ext></extLst>")
payload["xl/worksheets/sheet1.xml"] = payload["xl/worksheets/sheet1.xml"].replace(
    b"</worksheet>", EXT + b"</worksheet>")

with zipfile.ZipFile("extended.xlsx", "w", zipfile.ZIP_DEFLATED) as zout:
    for name, data in payload.items():
        zout.writestr(name, data)

EXTENSION_MARKERS = {
    "extLst": b"<extLst",
    "x14": b"x14:",
    "AlternateContent": b"mc:AlternateContent",
}


def scan_extension_markers(archive, info):
    found = set()
    overlap = max(len(marker) for marker in EXTENSION_MARKERS.values()) - 1
    tail = b""
    with archive.open(info) as stream:
        while chunk := stream.read(64 * 1024):
            window = tail + chunk
            found.update(label for label, marker in EXTENSION_MARKERS.items() if marker in window)
            if len(found) == len(EXTENSION_MARKERS):
                break
            tail = window[-overlap:]
    return found


def archive_inventory(source):
    with zipfile.ZipFile(source) as archive:
        names = set(archive.namelist())
        markers = {}
        for info in archive.infolist():
            if info.filename.endswith((".xml", ".rels")):
                found = scan_extension_markers(archive, info)
                if found:
                    markers[info.filename] = found
        return names, markers


def stripped_extension_markers(before, after, common_names):
    return sorted(
        (name, label)
        for name in common_names
        for label in before.get(name, set()) - after.get(name, set())
    )


def round_trip_changes(path, **load_options):
    before_names, before_markers = archive_inventory(path)
    wb = openpyxl.load_workbook(path, **load_options)  # same options as the real edit
    with TemporaryFile() as output:
        wb.save(output)
        output.seek(0)
        after_names, after_markers = archive_inventory(output)
    wb.close()
    dropped = sorted(before_names - after_names)
    stripped_extensions = stripped_extension_markers(
        before_markers, after_markers, before_names & after_names
    )
    return dropped, stripped_extensions


dropped, stripped = round_trip_changes("extended.xlsx")
check("injected slicer-like part is detected as dropped", "xl/slicers/slicer1.xml" in dropped, dropped)
check("in-part x14 extension strip is detected (same part name)",
      ("xl/worksheets/sheet1.xml", "x14") in stripped, stripped)
check("clean workbook reports nothing", round_trip_changes("plain.xlsx") == ([], []))
inventory_names, inventory_markers = archive_inventory("extended.xlsx")
check("binary archive parts are named but never marker-scanned",
      "xl/media/large.bin" in inventory_names and "xl/media/large.bin" not in inventory_markers)
partial_before = {"xl/worksheets/sheet1.xml": {"extLst", "x14"}}
partial_after = {"xl/worksheets/sheet1.xml": {"extLst"}}
check(
    "marker comparison catches x14 loss while extLst survives",
    stripped_extension_markers(partial_before, partial_after, set(partial_before))
    == [("xl/worksheets/sheet1.xml", "x14")],
    stripped_extension_markers(partial_before, partial_after, set(partial_before)),
)

# ---- formatting.md snippet: sheet references built from the real sheet title -------
wb_f = openpyxl.Workbook()
src = wb_f.active
src.title = "Raw Data"          # space forces quoting
src.append(["Region", "Units", "Price"])
src.append(["EU", 3, 10])
src.append(["US", 4, 20])
agg = wb_f.create_sheet("ByRegion")


def sheet_ref(sheet):
    escaped = sheet.title.replace("'", "''")
    return f"'{escaped}'!"


ref = sheet_ref(src)
agg.append(["Region", "Units"])
agg["A2"] = "EU"
agg["B2"] = f"=SUMIF({ref}A:A,A2,{ref}B:B)"
wb_f.save("agg.xlsx")
wb_g = openpyxl.load_workbook("agg.xlsx")
check("formula references the real sheet name", wb_g["ByRegion"]["B2"].value == "=SUMIF('Raw Data'!A:A,A2,'Raw Data'!B:B)", wb_g["ByRegion"]["B2"].value)
apostrophe_sheet = wb_f.create_sheet("O'Brien")
apostrophe_sheet["A1"] = 1
agg["B3"] = f"=SUM({sheet_ref(apostrophe_sheet)}A:A)"
wb_f.save("apostrophe-agg.xlsx")
apostrophe_formula = openpyxl.load_workbook("apostrophe-agg.xlsx")["ByRegion"]["B3"].value
check(
    "quoted sheet reference doubles apostrophes",
    apostrophe_formula == "=SUM('O''Brien'!A:A)",
    apostrophe_formula,
)
hyphen_sheet = wb_f.create_sheet("Q1-Data")
hyphen_sheet["A1"] = 1
agg["B4"] = f"=SUM({sheet_ref(hyphen_sheet)}A:A)"
wb_f.save("hyphen-agg.xlsx")
hyphen_formula = openpyxl.load_workbook("hyphen-agg.xlsx")["ByRegion"]["B4"].value
check("ambiguous punctuation is protected by quoting", hyphen_formula == "=SUM('Q1-Data'!A:A)", hyphen_formula)

# Falsey values are valid categories; blank filtering must not discard or conflate them.
falsey_ws = openpyxl.Workbook().active
for value in ("Category", 0, False, "", None, 0, False):
    falsey_ws.append([value])
regions = []
seen_region_keys = set()
for (region,) in falsey_ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    if region is None or region == "":
        continue
    key = (type(region), region)
    if key not in seen_region_keys:
        seen_region_keys.add(key)
        regions.append(region)
check(
    "aggregation preserves numeric zero and boolean false as distinct categories",
    len(regions) == 2
    and type(regions[0]) is int and regions[0] == 0
    and type(regions[1]) is bool and regions[1] is False,
    [(type(value).__name__, value) for value in regions],
)

# ---- edit.md structural audit includes non-cell dependencies ------------------
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.formula import Tokenizer
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table


def formula_text(value):
    if isinstance(value, str):
        return value
    if text := getattr(value, "text", None):
        return text
    fields = ("ref", "r1", "r2", "dt2D", "dtr", "ca", "del1", "del2")
    details = ", ".join(
        f"{name}={getattr(value, name)!r}" for name in fields if hasattr(value, name)
    )
    return f"{type(value).__name__}({details})"


def defined_name_values(workbook):
    names = workbook.defined_names
    return names.values() if hasattr(names, "values") else names.definedName


def drawing_anchor_rows(drawing):
    anchor = drawing.anchor
    if isinstance(anchor, str):
        return (coordinate_to_tuple(anchor)[0],)
    rows = []
    if marker := getattr(anchor, "_from", None):
        rows.append(marker.row + 1)
    if marker := getattr(anchor, "to", None):
        rows.append(marker.row + 1)
    return tuple(rows)


def structural_references(workbook):
    refs = []
    for item in defined_name_values(workbook):
        refs.append(("defined name", item.name, item.attr_text))
    for sheet in workbook.worksheets:
        owner = sheet.title
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    refs.append(("cell formula", f"{owner}!{cell.coordinate}",
                                 formula_text(cell.value)))
        for table in sheet.tables.values():
            refs.append(("table", owner + "!" + table.name, table.ref))
        for merged_range in sheet.merged_cells.ranges:
            refs.append(("merged range", owner, str(merged_range)))
        if sheet.auto_filter.ref:
            refs.append(("auto filter", owner, sheet.auto_filter.ref))
        for label, value in (
            ("print area", sheet.print_area),
            ("print title rows", sheet.print_title_rows),
            ("print title columns", sheet.print_title_cols),
        ):
            if value:
                refs.append((label, owner, str(value)))
        for validation in sheet.data_validations.dataValidation:
            refs.append(("data validation range", owner, str(validation.sqref)))
            for formula in (validation.formula1, validation.formula2):
                if formula:
                    refs.append(("data validation formula", owner, str(formula)))
        for conditional_range in sheet.conditional_formatting:
            refs.append(("conditional formatting range", owner, str(conditional_range.sqref)))
            for rule in sheet.conditional_formatting[conditional_range]:
                for formula in getattr(rule, "formula", ()):
                    refs.append(("conditional formatting formula", owner, str(formula)))
        for index, chart in enumerate(sheet._charts, start=1):
            refs.append(("drawing anchor", f"{owner} chart {index}", drawing_anchor_rows(chart)))
            for element in chart._write().iter():
                if element.tag.rsplit("}", 1)[-1] == "f" and element.text:
                    refs.append(("chart series", f"{owner} chart {index}", element.text))
        for index, image in enumerate(sheet._images, start=1):
            refs.append(("drawing anchor", f"{owner} image {index}", drawing_anchor_rows(image)))
    return refs


def cell_formula_references(workbook):
    refs = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    value = cell.value
                    refs.append((
                        "cell formula",
                        sheet.title,
                        cell.coordinate,
                        formula_text(value),
                    ))
    return refs


def formula_may_intersect_rows(owner_sheet, formula, shifted_sheet, start_row):
    if not isinstance(formula, str) or not formula.startswith("="):
        return True
    tokens = Tokenizer(formula).items
    unmodeled_reference_functions = {"indirect", "offset", "address"}
    if any(
        token.type == "FUNC" and token.subtype == "OPEN"
        and token.value.rstrip("(").rsplit(":", 1)[-1]
        .lstrip("@").rsplit(".", 1)[-1].casefold()
        in unmodeled_reference_functions
        for token in tokens
    ):
        return True
    for token in tokens:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        reference = token.value
        target_sheet = owner_sheet
        if "!" in reference:
            qualifier, reference = reference.rsplit("!", 1)
            if "[" in qualifier or ":" in qualifier:
                return True
            target_sheet = qualifier.strip("'").replace("''", "'")
        if target_sheet.casefold() != shifted_sheet.casefold():
            continue
        try:
            _, min_row, _, max_row = range_boundaries(reference.replace("$", ""))
        except ValueError:
            return True
        if min_row is None or max_row is None or max_row >= start_row:
            return True
    return False

class LegacyDefinedNames:
    """Minimal openpyxl 3.0-style DefinedNameList surface."""
    definedName = [DefinedName("LegacyRange", attr_text="'Legacy'!$A$1")]


class LegacyWorkbook:
    defined_names = LegacyDefinedNames()
    worksheets = []


check(
    "structural audit supports openpyxl 3.0 DefinedNameList",
    structural_references(LegacyWorkbook())
    == [("defined name", "LegacyRange", "'Legacy'!$A$1")],
)


audit_wb = openpyxl.Workbook()
audit_ws = audit_wb.active
audit_ws.title = "Audit"
audit_ws.append(["Value"])
audit_ws.append([1])
audit_ws.append([2])
audit_ws["C1"] = "=SUM(A2:A3)"
audit_wb.defined_names.add(DefinedName("AuditRange", attr_text="'Audit'!$A$2:$A$3"))
audit_ws.add_table(Table(displayName="AuditTable", ref="A1:A3"))
audit_ws.merge_cells("B2:B3")
audit_ws.auto_filter.ref = "A1:A3"
audit_ws.print_area = "A1:B3"
audit_ws.print_title_rows = "1:1"
audit_ws.print_title_cols = "A:A"
validation = DataValidation(type="whole", formula1="'Audit'!$A$2")
validation.add("A2:A3")
audit_ws.add_data_validation(validation)
audit_ws.conditional_formatting.add("A2:A3", FormulaRule(formula=["A2>0"]))
chart = BarChart()
chart.add_data(Reference(audit_ws, min_col=1, min_row=1, max_row=3), titles_from_data=True)
audit_ws.add_chart(chart, "C1")
with open("anchor.png", "wb") as stream:
    stream.write(base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9ZlRYAAAAASUVORK5CYII="
    ))
audit_ws.add_image(Image("anchor.png"), "E5")
audit_references = structural_references(audit_wb)
reference_kinds = {kind for kind, _, _ in audit_references}
formula_references = cell_formula_references(audit_wb)
check(
    "structural audit covers names, tables, filters, validation, formatting, and charts",
    {"defined name", "cell formula", "table", "merged range", "auto filter", "print area",
     "print title rows", "print title columns", "data validation range",
     "data validation formula", "conditional formatting range",
     "conditional formatting formula", "chart series", "drawing anchor"} <= reference_kinds,
    reference_kinds,
)
check("drawing audit records an image anchored at the insertion row",
      ("drawing anchor", "Audit image 1", (5,)) in audit_references, audit_references)
check(
    "structural audit snapshots ordinary cell formulas before row insertion",
    formula_references == [("cell formula", "Audit", "C1", "=SUM(A2:A3)")],
    formula_references,
)
check("intersecting formula ranges are blocked before row insertion",
      formula_may_intersect_rows("Audit", "=SUM(A2:A3)", "Audit", 3))
check("audited formula ranges above the insertion can proceed",
      not formula_may_intersect_rows("Data", "=C2*1.08", "Data", 5))
check("sheet qualifiers are matched case-insensitively",
      formula_may_intersect_rows("Summary", "=SUM(data!A5:A6)", "Data", 5))
check("a genuinely different sheet remains outside the shifted rows",
      not formula_may_intersect_rows("Summary", "=SUM(Archive!A5:A6)", "Data", 5))
check("INDIRECT string references require a manual structural rewrite plan",
      formula_may_intersect_rows("Data", '=SUM(INDIRECT("A5:A6"))', "Data", 5))
check("OFFSET numeric row references require a manual structural rewrite plan",
      formula_may_intersect_rows("Data", "=SUM(OFFSET(A1,4,0,2,1))", "Data", 5))
check("implicit-intersection INDIRECT references require a manual rewrite plan",
      formula_may_intersect_rows("Data", '=@INDIRECT("A5:A6")', "Data", 5))
check("OFFSET used by the range operator requires a manual rewrite plan",
      formula_may_intersect_rows("Data", "=SUM(A1:OFFSET(A1,5,0))", "Data", 5))
check("INDIRECT text inside a string does not create a dynamic reference",
      not formula_may_intersect_rows(
          "Data", '=IF(A1="INDIRECT(A5:A6)",1,0)', "Data", 5
      ))

stale_wb = openpyxl.Workbook()
stale_ws = stale_wb.active
stale_ws["A5"], stale_ws["A6"] = 10, 20
stale_ws["B1"] = "=SUM(A5:A6)"
stale_formula_before = cell_formula_references(stale_wb)
stale_ws.insert_rows(5)
check(
    "insert_rows leaves intersecting formulas stale (negative control)",
    stale_formula_before == [("cell formula", "Sheet", "B1", "=SUM(A5:A6)")]
    and stale_ws["B1"].value == "=SUM(A5:A6)"
    and (stale_ws["A6"].value, stale_ws["A7"].value) == (10, 20),
    (stale_formula_before, stale_ws["B1"].value),
)

safe_wb = openpyxl.Workbook()
safe_ws = safe_wb.active
safe_ws.title = "Data"
safe_ws["C2"], safe_ws["D2"] = 100, "=C2*1.08"
safe_before = cell_formula_references(safe_wb)
safe_dependencies = [
    reference for reference in safe_before
    if formula_may_intersect_rows(reference[1], reference[3], "Data", 5)
]
if not safe_dependencies:
    safe_ws.insert_rows(5)
    safe_wb.calculation.fullCalcOnLoad = True
    safe_wb.calculation.forceFullCalc = True
    safe_wb.calculation.calcMode = "auto"
    safe_wb.save("audited-structural-edit.xlsx")
safe_reopened = openpyxl.load_workbook("audited-structural-edit.xlsx", data_only=False)
check("audited non-intersecting formula path reaches save",
      safe_reopened["Data"]["D2"].value == "=C2*1.08")
check("formula edit forces recalculation after save",
      safe_reopened.calculation.fullCalcOnLoad is True
      and safe_reopened.calculation.calcMode == "auto")

legacy_name = DefinedName("LegacyName", attr_text="Audit!$A$1")
legacy_names = type("LegacyDefinedNames", (), {"definedName": [legacy_name]})()
legacy_workbook = type("LegacyWorkbook", (), {"defined_names": legacy_names})()
check("defined-name adapter supports openpyxl 3.0 collections",
      list(defined_name_values(legacy_workbook)) == [legacy_name])

formula_only = openpyxl.Workbook()
formula_only.active["A1"] = "=Data!A5"
check("structural guard sees formulas before row insertion",
      structural_references(formula_only)
      == [("cell formula", "Sheet!A1", "=Data!A5")],
      structural_references(formula_only))

audited_wb = openpyxl.Workbook()
audited_ws = audited_wb.active
audited_ws.title = "Data"
audited_ws["B2"] = "=A2*2"
audited_ws["A5"] = "shift me"
audited_references_before = structural_references(audited_wb)
audited_unchanged = {("cell formula", "Data!B2", "=A2*2")}
unaudited_references = [
    reference for reference in audited_references_before
    if reference not in audited_unchanged
]
if not unaudited_references:
    audited_ws.insert_rows(5)
    audited_ws["D2"] = "=C2*1.08"
    audited_wb.save("audited-edit.xlsx")
audited_reopened = openpyxl.load_workbook("audited-edit.xlsx", data_only=False)
check("exact dependency allowlist lets an audited edit reach save",
      audited_references_before == [("cell formula", "Data!B2", "=A2*2")]
      and unaudited_references == []
      and audited_reopened["Data"]["A6"].value == "shift me"
      and audited_reopened["Data"]["B2"].value == "=A2*2"
      and audited_reopened["Data"]["D2"].value == "=C2*1.08")

unsafe_wb = openpyxl.Workbook()
unsafe_ws = unsafe_wb.active
unsafe_ws.title = "Data"
unsafe_ws["A5"], unsafe_ws["A6"] = 10, 20
unsafe_ws["B1"] = "=SUM(A5:A6)"
unsafe_dependencies = [
    reference for reference in structural_references(unsafe_wb)
    if reference not in audited_unchanged
]
unsafe_blocked_before_insert = bool(unsafe_dependencies)
check("intersecting formula not on the exact allowlist is blocked before insertion",
      unsafe_blocked_before_insert
      and unsafe_ws["A5"].value == 10 and unsafe_ws["A6"].value == 20,
      unsafe_dependencies)

# and the edit itself still works after the warning path
wb2 = openpyxl.load_workbook("plain.xlsx")
wb2["Data"]["B2"] = "=B2*1"  # formula stays a formula
wb2["Data"]["B2"].number_format = "#,##0.00"
wb2.save("edited.xlsx")
wb3 = openpyxl.load_workbook("edited.xlsx")
check("edited cell keeps a formula string", isinstance(wb3["Data"]["B2"].value, str) and wb3["Data"]["B2"].value.startswith("="))
expected_number_formats = {"Data": {"B2": "#,##0.00"}}
format_matches = all(
    wb3[sheet][coordinate].number_format == expected
    for sheet, cells in expected_number_formats.items()
    for coordinate, expected in cells.items()
)
check("task-specific number format mapping is verified", format_matches)

# ---- read.md snippet: multi-sheet profiles cover every sheet ----------------------
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

wb_h = openpyxl.Workbook()
wb_h.active.title = "First"
wb_h.active["A1"] = "=1+1"
wb_h.active["A2"] = ArrayFormula("A2:A3", "=ROW(A2:A3)")
second = wb_h.create_sheet("Second")
second["A1"] = "plain"
second["A2"] = "=2+2"
second["B1"] = DataTableFormula(ref="B1:B2", r1="C1")
wb_h.save("multi.xlsx")
formula_wb = openpyxl.load_workbook("multi.xlsx", read_only=True, data_only=False)
value_wb = openpyxl.load_workbook("multi.xlsx", read_only=True, data_only=True)
profiled = list(value_wb.sheetnames)
uncached = {(sn, fc.coordinate, formula_text(fc.value))
            for sn in profiled
            for frow, vrow in zip(formula_wb[sn].iter_rows(), value_wb[sn].iter_rows())
            for fc, vc in zip(frow, vrow)
            if fc.data_type == "f" and vc.value is None}
check("multi-sheet profile iterates every sheet", profiled == ["First", "Second"], profiled)
check("uncached formulas found on both sheets", {item[0] for item in uncached} == {"First", "Second"}, uncached)
check("array-formula objects are detected by data_type", ("First", "A2", "=ROW(A2:A3)") in uncached, uncached)
data_table_entry = next(item for item in uncached if item[:2] == ("Second", "B1"))
check("data-table formulas have stable diagnostic text",
      data_table_entry[2].startswith("DataTableFormula(ref='B1:B2', r1='C1'")
      and "0x" not in data_table_entry[2], data_table_entry)

# csv.md: value export uses the cached-value workbook and reports every missing cache.
formula_wb["First"].reset_dimensions()
value_wb["First"].reset_dimensions()

def spreadsheet_safe_csv_value(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


missing_caches = []
with open("formula-values.csv", "w", newline="", encoding="utf-8") as output:
    writer = csv.writer(output)
    for formula_row, value_row in zip(formula_wb["First"].iter_rows(), value_wb["First"].iter_rows()):
        for formula_cell, value_cell in zip(formula_row, value_row):
            if formula_cell.data_type == "f" and value_cell.value is None:
                missing_caches.append(formula_cell.coordinate)
        writer.writerow([spreadsheet_safe_csv_value(cell.value) for cell in value_row])
with open("formula-values.csv", newline="", encoding="utf-8") as exported:
    exported_values = [value for row in csv.reader(exported) for value in row]
check("XLSX-to-CSV reports formulas with no cached value", set(missing_caches) >= {"A1", "A2"}, missing_caches)
check("XLSX-to-CSV does not leak formula strings into value output",
      not any(value.startswith("=") for value in exported_values), exported_values)

formula_like_literals = ["=1+1", "+cmd", "-2+3", "@SUM(A1:A2)", -5]
with open("formula-like-literals.csv", "w", newline="", encoding="utf-8") as output:
    csv.writer(output).writerow([
        spreadsheet_safe_csv_value(value) for value in formula_like_literals
    ])
with open("formula-like-literals.csv", newline="", encoding="utf-8") as exported:
    safe_literals = next(csv.reader(exported))
check("spreadsheet-targeted CSV neutralizes every formula-like text prefix",
      safe_literals[:4] == ["'=1+1", "'+cmd", "'-2+3", "'@SUM(A1:A2)"], safe_literals)
check("CSV neutralization leaves numeric negative values numeric",
      safe_literals[4] == "-5", safe_literals)
formula_wb.close()
value_wb.close()

# SKILL.md contract: fullCalcOnLoad makes viewers recalculate even in manual calc mode.
calc_wb = openpyxl.Workbook()
calc_ws = calc_wb.active
calc_ws["A1"] = 1
calc_ws["A2"] = 2
calc_ws["A3"] = "=SUM(A1:A2)"
calc_wb.calculation.calcMode = "manual"
calc_wb.calculation.fullCalcOnLoad = False  # simulate a source that does not recalc on load
calc_wb.save("stale-calc.xlsx")
stale_reopened = openpyxl.load_workbook("stale-calc.xlsx")
check("workbook without fullCalcOnLoad round-trips the stale flag (negative control)",
      not bool(getattr(stale_reopened.calculation, "fullCalcOnLoad", False)),
      stale_reopened.calculation)
stale_reopened.calculation.fullCalcOnLoad = True
stale_reopened.save("manual-calc.xlsx")
calc_reopened = openpyxl.load_workbook("manual-calc.xlsx")
check("fullCalcOnLoad survives save/reload",
      bool(getattr(calc_reopened.calculation, "fullCalcOnLoad", False)),
      calc_reopened.calculation)
check("manual calc mode survives save/reload",
      getattr(calc_reopened.calculation, "calcMode", None) == "manual",
      calc_reopened.calculation)
check("reloaded formula cell still holds the formula string",
      calc_reopened.active["A3"].value == "=SUM(A1:A2)", calc_reopened.active["A3"].value)
calc_reopened.close()


# ---- edit.md snippet: extension detection is prefix-independent ------------------
X14_URI = b"http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
MC_URI = b"http://schemas.openxmlformats.org/markup-compatibility/2006"
EXTENSION_MARKERS = {
    "extLst": b"extLst",
    "x14 namespace": X14_URI,
    "markup compatibility": MC_URI,
}

def markers_in(data):
    return {label for label, marker in EXTENSION_MARKERS.items() if marker in data}

custom_prefix_sheet = (
    b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
    b'xmlns:sx="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" '
    b'xmlns:ooo="http://schemas.openxmlformats.org/markup-compatibility/2006">'
    b'<ooo:AlternateContent><sx:extLst/></ooo:AlternateContent></worksheet>'
)
found = markers_in(custom_prefix_sheet)
check("namespace markers detect custom-prefix x14 extensions", "x14 namespace" in found, found)
check("namespace markers detect custom-prefix markup compatibility", "markup compatibility" in found, found)
check("local-name marker detects any-prefix extLst", "extLst" in found, found)
legacy_prefix_markers = {b"x14:", b"mc:AlternateContent"}
check("prefix markers are provably blind to custom prefixes (negative control)",
      not any(marker in custom_prefix_sheet for marker in legacy_prefix_markers))

# ---- formatting.md guard: header-only sheets skip conditional formatting ---------
def add_demo_formatting(sheet):
    last = next(
        (row for row in range(sheet.max_row, 1, -1)
         if any(sheet.cell(row, column).value is not None for column in range(1, 7))),
        1,
    )
    if last < 2:
        return 0
    sheet.conditional_formatting.add(
        f"D2:D{last}", CellIsRule(operator="lessThan", formula=["0"]),
    )
    sheet.conditional_formatting.add(
        f"A2:F{last}", FormulaRule(formula=["$D2<0"]),
    )
    sheet.conditional_formatting.add(
        f"C2:C{last}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="63BE7B"),
    )
    sheet.conditional_formatting.add(
        f"E2:E{last}", DataBarRule(start_type="min", end_type="max", color="638EC6"),
    )
    return 4


header_wb = openpyxl.Workbook()
header_ws = header_wb.active
header_ws.append(["A", "B", "C", "D", "E", "F"])
try:
    header_ws.conditional_formatting.add(
        "D2:D1", CellIsRule(operator="lessThan", formula=["0"]),
    )
    inverted_range_rejected = False
except (TypeError, ValueError):
    inverted_range_rejected = True
check("unguarded header-only range is rejected (negative control)", inverted_range_rejected)
check("header-only guard skips all four formatting rules", add_demo_formatting(header_ws) == 0)
header_wb.save("header-only-formatting.xlsx")
header_reopened = openpyxl.load_workbook("header-only-formatting.xlsx")
check("header-only workbook saves and reopens with no conditional formatting",
      len(header_reopened.active.conditional_formatting) == 0)

data_wb = openpyxl.Workbook()
data_ws = data_wb.active
data_ws.append(["A", "B", "C", "D", "E", "F"])
data_ws.append([1, 2, 3, -1, 5, 6])
data_ws["F10000"].number_format = "0.00"  # inflate max_row without adding data
check("data rows receive all four formatting rules", add_demo_formatting(data_ws) == 4)
formatting_ranges = [str(item.sqref) for item in data_ws.conditional_formatting]
check("conditional formatting ignores a styled ghost row",
      all("10000" not in item and item.endswith("2") for item in formatting_ranges),
      formatting_ranges)
data_wb.save("data-formatting.xlsx")
data_reopened = openpyxl.load_workbook("data-formatting.xlsx")
check("all four formatting rules survive save/reopen",
      len(data_reopened.active.conditional_formatting) == 4)


# ---- read.md: implausible <dimension> is reset before streaming ------------------
dim_wb = openpyxl.Workbook()
dim_ws = dim_wb.active
dim_ws.append(["h1", "h2"])
dim_ws.append([1, 2])
dim_ws.append([3, 4])
dim_ws["C4"] = "=SUM(A2:B3)"
dim_wb.save("dimension.xlsx")
# Corrupt the sheet's dimension metadata the way non-Excel producers do.
import zipfile as dim_zip
with dim_zip.ZipFile("dimension.xlsx") as archive:
    members = {name: archive.read(name) for name in archive.namelist()}
members["xl/worksheets/sheet1.xml"] = members["xl/worksheets/sheet1.xml"].replace(
    b"<dimension ref=\"A1:C4\"/>", b"<dimension ref=\"A1:B2\"/>"
)
with dim_zip.ZipFile("dimension.xlsx", "w") as archive:
    for name, data in members.items():
        archive.writestr(name, data)

from openpyxl.utils import get_column_letter


def discover_dimension(worksheet):
    worksheet.reset_dimensions()
    min_row = min_column = max_row = max_column = None
    for row in worksheet.iter_rows():
        for cell in row:
            row_index = getattr(cell, "row", None)
            column_index = getattr(cell, "column", None)
            if row_index is None or column_index is None:
                continue
            min_row = row_index if min_row is None else min(min_row, row_index)
            min_column = column_index if min_column is None else min(min_column, column_index)
            max_row = row_index if max_row is None else max(max_row, row_index)
            max_column = column_index if max_column is None else max(max_column, column_index)
    if max_row is None:
        return "A1:A1"
    return (f"{get_column_letter(min_column)}{min_row}:"
            f"{get_column_letter(max_column)}{max_row}")


dim_value = openpyxl.load_workbook("dimension.xlsx", read_only=True, data_only=True)
dim_formula = openpyxl.load_workbook("dimension.xlsx", read_only=True, data_only=False)
dim_ws_ro = dim_value.active
check("plausible but truncated dimension limits streaming (negative control)",
      dim_ws_ro.calculate_dimension() == "A1:B2" and dim_ws_ro.max_row == 2,
      (dim_ws_ro.calculate_dimension(), dim_ws_ro.max_row))
streamed_before_reset = [row for row in dim_ws_ro.iter_rows(values_only=True)]
discovered_value_dimension = discover_dimension(dim_ws_ro)
discovered_formula_dimension = discover_dimension(dim_formula.active)
streamed_after_reset = [row for row in dim_ws_ro.iter_rows(values_only=True)]
dim_value.close()
dim_formula.close()
check("dimension scan recovers plausible row/column truncation in both workbook views",
      discovered_value_dimension == "A1:C4"
      and discovered_formula_dimension == "A1:C4",
      (discovered_value_dimension, discovered_formula_dimension))
check("reset_dimensions restores the real extent",
      len(streamed_before_reset) == 2
      and len(streamed_after_reset) == 4
      and streamed_after_reset[2][:2] == (3, 4),
      (len(streamed_before_reset), streamed_after_reset[-1]))

empty_dimension_wb = openpyxl.Workbook()
empty_dimension_wb.save("empty-dimension.xlsx")
empty_dimension_ro = openpyxl.load_workbook("empty-dimension.xlsx", read_only=True)
check("dimension scan handles an actually empty worksheet",
      discover_dimension(empty_dimension_ro.active) == "A1:A1")
empty_dimension_ro.close()


print("\n" + ("ALL XLSX FIXTURES PASSED" if not failures else f"{len(failures)} FAILURES: {failures}"))
sys.exit(0 if not failures else 1)
